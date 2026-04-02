# -*- coding: utf-8 -*-
"""
Natural Gas Simple Flowsheet
=============================
Python translation of NaturalGasSimpleFlowsheet.Generate() (C#).

Flowsheet: NG feed -> COM-01 -> SPLIT-01 -> (bypass / coolers) -> MIX-01
           -> HEX-01 -> COOL-03 -> FLASH-01 -> SPLIT-02
           -> (HEX-02 / EXP-01 path) + VALV-01 + VALV-02
           -> MIX-DEMETH -> DEMETH (ComponentSeparator)
                         -> Methane product (S-24)
                         -> S-25 -> DEETH (ComponentSeparator)
                                 -> Ethane product
                                 -> Deeth_Btm product

Adaptations vs. original C#:
  - Vessel energy input: ConEn connector is at port 6, not 1
  - 4-inlet DEMETH: routed through MIX-DEMETH (ComponentSeparator only has 1 inlet)
  - S-24 -> Methane stream-to-stream is skipped (invalid); S-24 IS the methane product
  - Splitter ratios set via .NET reflection on the Ratios ArrayList property
  - ComponentSepSpecs set via .NET reflection on the concrete ComponentSeparator type

Rule: if any unit-op ADD raises a DLL / type-load error -> stop immediately.
"""

import os
import sys
import logging
from datetime import datetime

project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

logging.basicConfig(level=logging.WARNING, format="%(levelname)s | %(message)s")

from src.core.automation import DWSIMAutomation
from src.core.flowsheet import FlowsheetManager

OUTPUT_FILE  = os.path.join(project_root, "outputs", "natural_gas_simple.dwxml")
EXCEL_FILE   = os.path.join(project_root, "outputs", "natural_gas_results.xlsx")

# Global connection log (src, src_port, tgt, tgt_port, connected?)
_CONNECTIONS: list[dict] = []

# ── Helpers ─────────────────────────────────────────────────────────────────────

def die(msg: str) -> None:
    print(f"\n[STOP] {msg}")
    sys.exit(1)


_DLL_KEYWORDS = ("typeloa", "filenotfound", "badimage", "assembly", ".dll",
                 "could not load", "unable to load")


def add_unitop(mgr: FlowsheetManager, type_key: str, name: str,
               x: float, y: float) -> object:
    """Add a unit-op and stop if a DLL/type-load error is raised."""
    try:
        obj = mgr.add_object(type_key, name, x, y)
        if obj is None:
            die(f"DLL/type error creating '{name}' ({type_key}): AddObject returned None")
        print(f"  [OK] {name} ({type_key})")
        return obj
    except Exception as exc:
        msg = str(exc).lower()
        if any(k in msg for k in _DLL_KEYWORDS):
            die(f"DLL error creating '{name}' ({type_key}): {exc}")
        die(f"Error creating '{name}' ({type_key}): {exc}")


def _set(obj, name: str, value) -> None:
    """Set a .NET attribute; warn on failure."""
    try:
        setattr(obj, name, value)
    except Exception as exc:
        print(f"  WARNING: could not set {name} = {value!r}: {exc}")


def _set_first(obj, *name_value_pairs) -> None:
    """Try (attr, value) pairs in order; stop at first success."""
    for name, value in name_value_pairs:
        try:
            setattr(obj, name, value)
            return
        except Exception:
            pass
    names = [n for n, _ in name_value_pairs]
    print(f"  WARNING: none of {names} could be set on {obj}")


def conn(mgr: FlowsheetManager, src: str, tgt: str,
         sp: int = 0, tp: int = 0) -> None:
    ok = mgr.connect(src, tgt, sp, tp)
    _CONNECTIONS.append({"src": src, "src_port": sp,
                         "tgt": tgt, "tgt_port": tp, "ok": ok})
    if not ok:
        print(f"  WARNING: could not connect {src} -> {tgt} [{sp}:{tp}]")


# ── Initialise DWSIM ─────────────────────────────────────────────────────────────

print("=" * 60)
print("Natural Gas Simple Flowsheet")
print("=" * 60)

print("\n[1] Initialising DWSIM...")
dwsim = DWSIMAutomation()
dwsim.initialize()

# .NET imports must come AFTER initialize()
import System
from System.Reflection import BindingFlags
_PUB_INST = BindingFlags.Public | BindingFlags.Instance


def _refl_get(obj, prop_name):
    """Return value of a property via .NET reflection (bypasses interface typing)."""
    prop = obj.GetType().GetProperty(prop_name, _PUB_INST)
    if prop is None:
        raise AttributeError(f"Property '{prop_name}' not found on {obj.GetType().FullName}")
    return prop.GetValue(obj, None)


def _refl_set(obj, prop_name, value) -> None:
    """Set a property via .NET reflection."""
    prop = obj.GetType().GetProperty(prop_name, _PUB_INST)
    if prop is None:
        raise AttributeError(f"Property '{prop_name}' not found on {obj.GetType().FullName}")
    prop.SetValue(obj, value, None)


fs = dwsim.create_flowsheet()
try:
    fs.Options.SimulationName = "Natural Gas Separation - Simple to Deethanizer"
except Exception:
    pass  # Options may not be exposed by all Automation3 builds
mgr = FlowsheetManager(fs, dwsim)

# ── Compounds ───────────────────────────────────────────────────────────────────

print("\n[2] Adding compounds...")
COMPOUNDS = ["Methane", "Ethane", "Propane", "Isobutane", "n-Butane",
             "Isopentane", "n-Pentane", "n-Hexane", "n-Heptane", "Nitrogen"]
for c in COMPOUNDS:
    mgr.add_compound(c)
    print(f"  + {c}")

# ── Property package ─────────────────────────────────────────────────────────────

print("\n[3] Setting Peng-Robinson property package...")
mgr.set_property_package("Peng-Robinson")

# ── Material streams ─────────────────────────────────────────────────────────────

print("\n[4] Adding material streams...")
s_feed    = mgr.add_object("material_stream", "NG-Feed",     50, 120)
s02       = mgr.add_object("material_stream", "S-02",       150, 120)
s03       = mgr.add_object("material_stream", "S-03",       250,  60)
s04       = mgr.add_object("material_stream", "S-04",       250, 150)
s05       = mgr.add_object("material_stream", "S-05",       350, 150)
s06       = mgr.add_object("material_stream", "S-06",       450, 150)
s07       = mgr.add_object("material_stream", "S-07",       450,  90)
s08       = mgr.add_object("material_stream", "S-08",       600,  90)
s10       = mgr.add_object("material_stream", "S-10",       700,  10)
s11       = mgr.add_object("material_stream", "S-11",       750,  90)
s12       = mgr.add_object("material_stream", "S-12",       900,  90)
s13       = mgr.add_object("material_stream", "S-13",       900, 180)
s14       = mgr.add_object("material_stream", "S-14",      1020, 120)
s15       = mgr.add_object("material_stream", "S-15",      1020,  60)
s16       = mgr.add_object("material_stream", "S-16",      1180,  30)
s17       = mgr.add_object("material_stream", "S-17",      1180,  70)
s19       = mgr.add_object("material_stream", "S-19",      1180, 140)
s20       = mgr.add_object("material_stream", "S-20",      1330,  80)
s21       = mgr.add_object("material_stream", "S-21",      1330, 140)
s22       = mgr.add_object("material_stream", "S-22",      1330, 190)
s23       = mgr.add_object("material_stream", "S-23",      1100, 210)
s_dmin    = mgr.add_object("material_stream", "S-DEMETH-IN",1400, 130)  # mixer -> DEMETH
s24       = mgr.add_object("material_stream", "S-24",      1520, 120)
s25       = mgr.add_object("material_stream", "S-25",      1520, 190)
s_eth     = mgr.add_object("material_stream", "Ethane",    1850, 160)
s_dbtm    = mgr.add_object("material_stream", "Deeth_Btm", 1850, 240)
# Cold-side utility inlets for HEX-01 and HEX-02
# (original C# code omitted these; required by DWSIM for 2-stream heat exchanger)
s_hex1c   = mgr.add_object("material_stream", "S-HEX1-COLD-IN",  640, 200)
s_hex2c   = mgr.add_object("material_stream", "S-HEX2-COLD-IN", 1110, 200)
print(f"  Added {28} material streams (incl. S-DEMETH-IN and cold-side HEX feeds)")

# ── Energy streams ───────────────────────────────────────────────────────────────

print("\n[5] Adding energy streams...")
e1  = mgr.add_object("energy_stream", "E-1",   120, 180)
e02 = mgr.add_object("energy_stream", "E-02",  330, 220)
e03 = mgr.add_object("energy_stream", "E-03",  430, 220)
e04 = mgr.add_object("energy_stream", "E-04",  720, 220)
e05 = mgr.add_object("energy_stream", "E-05",  860, 240)
e08 = mgr.add_object("energy_stream", "E-08", 1280, 240)
print(f"  Added 6 energy streams")

# ── Unit operations (DLL-checked) ────────────────────────────────────────────────

print("\n[6] Adding unit operations (DLL check active)...")
com01     = add_unitop(mgr, "compressor",     "COM-01",     100, 120)
split01   = add_unitop(mgr, "splitter",       "SPLIT-01",   200, 120)
cool01    = add_unitop(mgr, "cooler",         "COOL-01",    300, 150)
cool02    = add_unitop(mgr, "cooler",         "COOL-02",    400, 150)
mix01     = add_unitop(mgr, "mixer",          "MIX-01",     500,  90)
hex01     = add_unitop(mgr, "heat_exchanger", "HEX-01",     650,  90)
cool03    = add_unitop(mgr, "cooler",         "COOL-03",    720,  90)
flash01   = add_unitop(mgr, "flash",          "FLASH-01",   840, 100)
split02   = add_unitop(mgr, "splitter",       "SPLIT-02",   970,  90)
hex02     = add_unitop(mgr, "heat_exchanger", "HEX-02",    1120,  60)
exp01     = add_unitop(mgr, "expander",       "EXP-01",    1120, 130)
flash02   = add_unitop(mgr, "flash",          "FLASH-02",  1260, 140)
valv01    = add_unitop(mgr, "valve",          "VALV-01",   1030, 210)
valv02    = add_unitop(mgr, "valve",          "VALV-02",   1260,  80)
mix_deth  = add_unitop(mgr, "mixer",          "MIX-DEMETH",1380, 150)  # combines 4 feeds
demeth    = add_unitop(mgr, "separator",      "DEMETH",    1450, 150)
deeth     = add_unitop(mgr, "separator",      "DEETH",     1780, 200)

# ── Connections ──────────────────────────────────────────────────────────────────

print("\n[7] Connecting objects...")

# Feed -> Compressor
conn(mgr, "NG-Feed",  "COM-01",   0, 0)   # material inlet
conn(mgr, "E-1",      "COM-01",   0, 1)   # energy  inlet

# Compressor -> Splitter
conn(mgr, "COM-01",   "S-02",     0, 0)
conn(mgr, "S-02",     "SPLIT-01", 0, 0)
conn(mgr, "SPLIT-01", "S-03",     0, 0)   # bypass fraction -> mixer
conn(mgr, "SPLIT-01", "S-04",     1, 0)   # main fraction   -> coolers

# Coolers path -> Mixer-01
conn(mgr, "S-03",     "MIX-01",   0, 0)   # bypass inlet
conn(mgr, "S-04",     "COOL-01",  0, 0)
conn(mgr, "E-02",     "COOL-01",  0, 1)   # energy
conn(mgr, "COOL-01",  "S-05",     0, 0)
conn(mgr, "S-05",     "COOL-02",  0, 0)
conn(mgr, "E-03",     "COOL-02",  0, 1)   # energy
conn(mgr, "COOL-02",  "S-06",     0, 0)
conn(mgr, "S-06",     "MIX-01",   0, 1)   # cooled stream inlet

# Mixer-01 -> HEX-01 -> COOL-03 -> FLASH-01
conn(mgr, "MIX-01",   "S-07",     0, 0)
conn(mgr, "S-07",           "HEX-01",   0, 0)  # hot side inlet
conn(mgr, "S-HEX1-COLD-IN","HEX-01",   0, 1)  # cold side inlet (utility stream)
conn(mgr, "HEX-01",   "S-08",     0, 0)   # hot side outlet -> COOL-03
conn(mgr, "HEX-01",   "S-10",     1, 0)   # cold side outlet (utility warmed up)
conn(mgr, "S-08",     "COOL-03",  0, 0)
conn(mgr, "E-04",     "COOL-03",  0, 1)   # energy
conn(mgr, "COOL-03",  "S-11",     0, 0)
conn(mgr, "S-11",     "FLASH-01", 0, 0)   # material inlet to flash
# NOTE: Vessel ConEn connector is at physical port index 6 (not 1 as in original C#)
conn(mgr, "E-05",     "FLASH-01", 0, 6)   # energy  inlet  (ConEn = index 6)

# FLASH-01 -> SPLIT-02 (vapour) + VALV-01 (liquid)
conn(mgr, "FLASH-01", "S-12",     0, 0)   # vapour outlet
conn(mgr, "FLASH-01", "S-13",     1, 0)   # liquid outlet
conn(mgr, "S-12",     "SPLIT-02", 0, 0)
conn(mgr, "SPLIT-02", "S-15",     0, 0)   # fraction -> HEX-02
conn(mgr, "SPLIT-02", "S-14",     1, 0)   # fraction -> EXP-01

# HEX-02 path -> VALV-02 -> MIX-DEMETH
conn(mgr, "S-15",           "HEX-02",   0, 0)  # hot side inlet
conn(mgr, "S-HEX2-COLD-IN","HEX-02",   0, 1)  # cold side inlet (utility stream)
conn(mgr, "HEX-02",   "S-17",     0, 0)   # hot side outlet
conn(mgr, "HEX-02",   "S-16",     1, 0)   # cold side outlet (utility warmed up)
conn(mgr, "S-17",     "VALV-02",  0, 0)
conn(mgr, "VALV-02",  "S-20",     0, 0)

# EXP-01 -> FLASH-02 -> MIX-DEMETH
conn(mgr, "S-14",     "EXP-01",   0, 0)
conn(mgr, "EXP-01",   "S-19",     0, 0)
conn(mgr, "S-19",     "FLASH-02", 0, 0)
# NOTE: same energy port fix applies to FLASH-02
conn(mgr, "E-08",     "FLASH-02", 0, 6)   # energy inlet (ConEn = index 6)
conn(mgr, "FLASH-02", "S-21",     0, 0)   # vapour outlet
conn(mgr, "FLASH-02", "S-22",     1, 0)   # liquid outlet

# VALV-01 -> MIX-DEMETH (liquid from FLASH-01)
conn(mgr, "S-13",     "VALV-01",  0, 0)
conn(mgr, "VALV-01",  "S-23",     0, 0)

# MIX-DEMETH: combine S-20, S-21, S-22, S-23 -> S-DEMETH-IN -> DEMETH
# (ComponentSeparator has only 1 material inlet; the mixer merges the 4 streams)
conn(mgr, "S-20",     "MIX-DEMETH", 0, 0)
conn(mgr, "S-21",     "MIX-DEMETH", 0, 1)
conn(mgr, "S-22",     "MIX-DEMETH", 0, 2)
conn(mgr, "S-23",     "MIX-DEMETH", 0, 3)
conn(mgr, "MIX-DEMETH", "S-DEMETH-IN", 0, 0)
conn(mgr, "S-DEMETH-IN", "DEMETH",   0, 0)

# DEMETH -> S-24 (Methane product) + S-25 -> DEETH
conn(mgr, "DEMETH",   "S-24",     0, 0)   # overhead
conn(mgr, "DEMETH",   "S-25",     1, 0)   # bottoms -> DEETH
# NOTE: S-24 -> Methane stream-to-stream is invalid in DWSIM; S-24 IS the product

# DEETH -> Ethane product + Deeth_Btm bottoms
conn(mgr, "S-25",     "DEETH",    0, 0)
conn(mgr, "DEETH",    "Ethane",   0, 0)   # overhead
conn(mgr, "DEETH",    "Deeth_Btm",1, 0)  # bottoms

print("  All connections applied.")

# ── Configure feed ────────────────────────────────────────────────────────────────

print("\n[8] Configuring feed stream...")
s_feed.set_conditions(
    temperature_K    = 310.0,
    pressure_Pa      = 59.7818e5,         # 59.7818 bar
    molar_flow_mol_s = 1464.0 / 3600.0,   # mol/s
    composition      = {
        "Methane":    0.8640,
        "Ethane":     0.0647,
        "Propane":    0.0287,
        "Isobutane":  0.0072,
        "n-Butane":   0.0082,
        "Isopentane": 0.0041,
        "n-Pentane":  0.0031,
        "n-Hexane":   0.0031,
        "n-Heptane":  0.0015,
        "Nitrogen":   0.0154,
    },
)
print("  Feed: T=310 K, P=59.78 bar, F=1464 mol/h")

# Cold-side utility streams for HEX-01 and HEX-02
# Pure methane at cryogenic conditions — acts as the "refrigerant" side of
# the cold-box exchangers that the C# code did not explicitly connect.
COLD_COMP = {c: (1.0 if c == "Methane" else 0.0) for c in COMPOUNDS}
s_hex1c.set_conditions(temperature_K=200.0, pressure_Pa=60e5,
                       molar_flow_mol_s=1000.0 / 3600.0,
                       composition=COLD_COMP)
s_hex2c.set_conditions(temperature_K=200.0, pressure_Pa=60e5,
                       molar_flow_mol_s=1000.0 / 3600.0,
                       composition=COLD_COMP)
print("  Cold-side HEX feeds: T=200 K, P=60 bar, F=1000 mol/h, pure CH4")

# ── Configure unit operations ─────────────────────────────────────────────────────

print("\n[9] Configuring unit operations...")

# ---- COM-01 -------------------------------------------------------------------
_set(com01, "CalcMode", 1)           # 1 = Delta_P
_set(com01, "DeltaP", 202650.0)
_set_first(com01,
           ("AdiabaticEfficiency", 0.75),
           ("Efficiency",          0.75))
print("  COM-01: CalcMode=DeltaP, dP=202650 Pa, eff=0.75")

# ---- SPLIT-01 via reflection on Ratios ArrayList  ----------------------------
# Ratios ArrayList: index 0 = outlet-0 fraction, index 1 = outlet-1 fraction
try:
    ratios = _refl_get(split01, "Ratios")
    ratios[0] = 15.83 / 1464.0
    ratios[1] = 1448.0 / 1464.0
    print(f"  SPLIT-01: ratios {15.83/1464:.5f} / {1448/1464:.5f} (bypass / main)")
except Exception as exc:
    print(f"  WARNING: SPLIT-01 Ratios: {exc}")

# ---- COOL-01 ------------------------------------------------------------------
_set(cool01, "CalcMode", 0)          # 0 = OutletTemperature
_set(cool01, "OutletTemperature", 305.3)
_set(cool01, "DeltaP", 20000.0)
print("  COOL-01: Tout=305.3 K, dP=20000 Pa")

# ---- COOL-02 ------------------------------------------------------------------
_set(cool02, "CalcMode", 0)
_set(cool02, "OutletTemperature", 289.7)
_set(cool02, "DeltaP", 20000.0)
print("  COOL-02: Tout=289.7 K, dP=20000 Pa")

# ---- HEX-01 (SpecifyUA mode = 0, UA = U × A) ----------------------------------
# Using SPECIFY_UA (mode 0) so DWSIM uses both inlet streams to solve.
# Original C# set U=1000 W/(m2·K) and A=239.2836 m2 → UA = 239 284 W/K
_set(hex01, "CalculationMode", 0)
_set(hex01, "OverallUA", 1000.0 * 239.2836)   # 239 284 W/K
print("  HEX-01: CalcMode=SpecifyUA, UA=239284 W/K")

# ---- COOL-03 ------------------------------------------------------------------
_set(cool03, "CalcMode", 0)
_set(cool03, "OutletTemperature", 251.4)
_set(cool03, "DeltaP", 0.0)
print("  COOL-03: Tout=251.4 K, dP=0")

# ---- FLASH-01 (PressureCalculation = Minimum = 1) ----------------------------
_set(flash01, "PressureCalculation", 1)
print("  FLASH-01: PressureCalculation=Minimum")

# ---- SPLIT-02 ----------------------------------------------------------------
try:
    ratios2 = _refl_get(split02, "Ratios")
    ratios2[0] = 483.3 / 1381.0
    ratios2[1] = 898.0 / 1381.0
    print(f"  SPLIT-02: ratios {483.3/1381:.5f} / {898/1381:.5f}")
except Exception as exc:
    print(f"  WARNING: SPLIT-02 Ratios: {exc}")

# ---- HEX-02 (SpecifyUA mode = 0, UA = U × A) ----------------------------------
# UA = 1000 × 804.7164 = 804 716 W/K
_set(hex02, "CalculationMode", 0)
_set(hex02, "OverallUA", 1000.0 * 804.7164)   # 804 716 W/K
print("  HEX-02: CalcMode=SpecifyUA, UA=804716 W/K")

# ---- EXP-01 (Delta_P mode = 1) -----------------------------------------------
_set(exp01, "CalcMode", 1)
_set(exp01, "DeltaP", 3577572.5)
_set_first(exp01,
           ("AdiabaticEfficiency", 0.75),
           ("Efficiency",          0.75))
print("  EXP-01: CalcMode=DeltaP, dP=3577573 Pa, eff=0.75")

# ---- FLASH-02 ----------------------------------------------------------------
_set(flash02, "PressureCalculation", 1)
print("  FLASH-02: PressureCalculation=Minimum")

# ---- VALV-01 -----------------------------------------------------------------
_set(valv01, "CalcMode", 0)          # 0 = OutletPressure
_set(valv01, "OutletPressure", 26.4133e5)
print("  VALV-01: Pout=26.4133 bar")

# ---- VALV-02 -----------------------------------------------------------------
_set(valv02, "CalcMode", 0)
_set(valv02, "OutletPressure", 27.0083e5)
print("  VALV-02: Pout=27.0083 bar")

# ---- DEMETH / DEETH: ComponentSepSpecs via reflection ------------------------

def configure_component_separator(sep_obj, tag: str, specs_pct: dict) -> None:
    """
    Configure a ComponentSeparator using PercentInletMolarFlow specs.
    Each value in specs_pct is the % of that component that goes to outlet-0
    (SpecifiedStreamIndex = 0).

    Uses .NET reflection to access ComponentSepSpecs on the concrete type
    (ISimulationObject doesn't expose it directly).
    """
    # SpecifiedStreamIndex is on the concrete type but often accessible via setattr
    _set(sep_obj, "SpecifiedStreamIndex", System.Byte(0))

    from DWSIM.UnitOperations.UnitOperations.Auxiliary import (
        ComponentSeparationSpec,
        SeparationSpec,
    )

    try:
        sep_specs = _refl_get(sep_obj, "ComponentSepSpecs")
        # sep_specs is Dictionary<string, ComponentSeparationSpec>
        # SeparationSpec.PercentInletMolarFlow = 3
        for comp, pct in specs_pct.items():
            spec = ComponentSeparationSpec(
                comp,
                SeparationSpec.PercentInletMolarFlow,
                float(pct),
                "",
            )
            sep_specs[comp] = spec
        print(f"  {tag}: ComponentSepSpecs set for {len(specs_pct)} components "
              f"(PercentInletMolarFlow)")
    except Exception as exc:
        print(f"  WARNING: {tag} ComponentSepSpecs via reflection failed: {exc}")
        # Fallback: use SplitFactors if available
        try:
            sf = _refl_get(sep_obj, "SplitFactors")
            for comp, pct in specs_pct.items():
                sf[comp] = float(pct) / 100.0
            print(f"  {tag}: SplitFactors fallback applied")
        except Exception as exc2:
            print(f"  WARNING: {tag} SplitFactors fallback also failed: {exc2}")


configure_component_separator(demeth, "DEMETH", {
    "Methane":    99.5,
    "Nitrogen":   99.5,
    "Ethane":      8.0,
    "Propane":     1.0,
    "Isobutane":   0.1,
    "n-Butane":    0.1,
    "Isopentane":  0.01,
    "n-Pentane":   0.01,
    "n-Hexane":    0.01,
    "n-Heptane":   0.01,
})

configure_component_separator(deeth, "DEETH", {
    "Methane":    99.0,
    "Ethane":     96.0,
    "Propane":     5.0,
    "Isobutane":   0.1,
    "n-Butane":    0.1,
    "Isopentane":  0.01,
    "n-Pentane":   0.01,
    "n-Hexane":    0.01,
    "n-Heptane":   0.01,
    "Nitrogen":   95.0,
})

# ══════════════════════════════════════════════════════════════════════════════
# Helper: collect stream results for the Excel / dynamic study
# ══════════════════════════════════════════════════════════════════════════════

ALL_STREAMS = [
    "NG-Feed", "S-02", "S-03", "S-04", "S-05", "S-06", "S-07", "S-08",
    "S-10", "S-11", "S-12", "S-13", "S-14", "S-15", "S-16", "S-17",
    "S-19", "S-20", "S-21", "S-22", "S-23", "S-DEMETH-IN",
    "S-24", "S-25", "Ethane", "Deeth_Btm",
]
PRODUCT_STREAMS = ["S-24", "Ethane", "Deeth_Btm"]


def collect_stream_results(mgr: FlowsheetManager) -> dict:
    """Return {stream_name: {T_K, P_bar, F_molh, VF, comp{}}} for all streams."""
    out = {}
    for sname in ALL_STREAMS:
        res = mgr.get_stream_results(sname)
        if res is None:
            out[sname] = None
            continue
        out[sname] = {
            "T_K":    res.get("temperature_K",    0.0),
            "P_bar":  res.get("pressure_Pa",       0.0) / 1e5,
            "F_molh": res.get("molar_flow_mol_s",  0.0) * 3600.0,
            "VF":     res.get("vapor_fraction",    0.0),
            "comp":   res.get("composition", {}).get("overall", {}),
        }
    return out


# ══════════════════════════════════════════════════════════════════════════════
# [10]  BASE-CASE CALCULATION
# ══════════════════════════════════════════════════════════════════════════════

print("\n[10] Calculating base-case flowsheet...")
result = dwsim.calculate(fs)

if result["errors"]:
    print(f"  Completed with {len(result['errors'])} error(s):")
    for err in result["errors"]:
        print(f"    - {err}")
else:
    print("  Calculation completed successfully.")

base_case = collect_stream_results(mgr)

# ── Console results ─────────────────────────────────────────────────────────
print("\n[11] Base-case stream results:")
print(f"  {'Stream':<16}  {'T (K)':>8}  {'T (C)':>7}  {'P (bar)':>8}"
      f"  {'F (mol/h)':>10}  {'VF':>5}")
print("  " + "-" * 68)
for sname in ALL_STREAMS:
    r = base_case.get(sname)
    if r is None:
        continue
    T_K, P, F, VF = r["T_K"], r["P_bar"], r["F_molh"], r["VF"]
    if T_K < 50:          # uncalculated default
        continue
    print(f"  {sname:<16}  {T_K:8.2f}  {T_K-273.15:7.2f}  {P:8.3f}"
          f"  {F:10.2f}  {VF:5.3f}")

print("\n  --- Product summary ---")
PROD_LABELS = {"S-24": "Methane (S-24)",
               "Ethane":    "Ethane product",
               "Deeth_Btm": "C3+ bottoms"}
for sname, lbl in PROD_LABELS.items():
    r = base_case.get(sname)
    if not r:
        continue
    comp = r["comp"]
    ch4  = comp.get("Methane",  0.0) * 100
    c2h6 = comp.get("Ethane",   0.0) * 100
    c3h8 = comp.get("Propane",  0.0) * 100
    print(f"  {lbl:<20}: F={r['F_molh']:7.2f} mol/h  "
          f"CH4={ch4:5.2f}%  C2H6={c2h6:5.2f}%  C3H8={c3h8:5.2f}%")

# ══════════════════════════════════════════════════════════════════════════════
# [12]  DYNAMIC SIMULATION — PARAMETRIC STUDIES
# ══════════════════════════════════════════════════════════════════════════════

print("\n[12] Running dynamic / parametric studies...")

NOMINAL_T    = 310.0
NOMINAL_P    = 59.7818e5
NOMINAL_F    = 1464.0 / 3600.0
NOMINAL_COMP = {
    "Methane":    0.8640, "Ethane":     0.0647, "Propane":    0.0287,
    "Isobutane":  0.0072, "n-Butane":   0.0082, "Isopentane": 0.0041,
    "n-Pentane":  0.0031, "n-Hexane":   0.0031, "n-Heptane":  0.0015,
    "Nitrogen":   0.0154,
}
# Non-methane baseline fractions (used for composition scaling)
_NON_CH4 = {k: v for k, v in NOMINAL_COMP.items() if k != "Methane"}
_NON_CH4_SUM = sum(_NON_CH4.values())   # 0.1360


def scaled_comp(c2plus_scale: float) -> dict:
    """Return composition where non-CH4 fractions are scaled by c2plus_scale,
    methane fills the remainder to 1.0."""
    scaled = {k: v * c2plus_scale for k, v in _NON_CH4.items()}
    scaled_sum = sum(scaled.values())
    scaled["Methane"] = max(0.0, 1.0 - scaled_sum)
    return scaled


def run_case(T_K: float, P_Pa: float, F_mol_s: float,
             comp: dict) -> dict:
    """Modify feed, re-calculate, return stream results dict."""
    s_feed.set_conditions(temperature_K=T_K, pressure_Pa=P_Pa,
                          molar_flow_mol_s=F_mol_s, composition=comp)
    dwsim.calculate(fs)
    return collect_stream_results(mgr)


def product_kpis(results: dict) -> dict:
    """Extract key product KPIs from a results dict."""
    kpis = {}
    for sname in ["S-24", "Ethane", "Deeth_Btm"]:
        r = results.get(sname)
        if r:
            kpis[sname] = {
                "F_molh": r["F_molh"],
                "CH4_pct":  r["comp"].get("Methane", 0) * 100,
                "C2H6_pct": r["comp"].get("Ethane",  0) * 100,
                "C3H8_pct": r["comp"].get("Propane", 0) * 100,
                "VF":       r["VF"],
            }
    return kpis


# ── Study A: Feed Temperature ────────────────────────────────────────────────
print("  Study A: Feed temperature sweep (P=59.78 bar, nominal comp)")
TEMP_CASES = [280.0, 295.0, 310.0, 325.0, 340.0]
study_A = []
for T in TEMP_CASES:
    r = run_case(T, NOMINAL_P, NOMINAL_F, NOMINAL_COMP)
    kpis = product_kpis(r)
    study_A.append({"T_K": T, "results": r, "kpis": kpis})
    feed_vf = r["NG-Feed"]["VF"] if r.get("NG-Feed") else 0
    ch4_f   = kpis.get("S-24",      {}).get("F_molh", 0)
    eth_f   = kpis.get("Ethane",    {}).get("F_molh", 0)
    btm_f   = kpis.get("Deeth_Btm", {}).get("F_molh", 0)
    print(f"    T={T:.0f} K ({T-273.15:+.1f} C): "
          f"VF_feed={feed_vf:.3f}  S-24={ch4_f:.1f}  Eth={eth_f:.1f}  Btm={btm_f:.1f} mol/h")

# ── Study B: Feed Pressure ───────────────────────────────────────────────────
print("  Study B: Feed pressure sweep (T=310 K, nominal comp)")
PRES_CASES = [50.0e5, 54.89e5, 59.78e5, 64.68e5, 69.57e5]
PRES_LABELS = [50.0, 54.89, 59.78, 64.68, 69.57]
study_B = []
for P, Plbl in zip(PRES_CASES, PRES_LABELS):
    r = run_case(NOMINAL_T, P, NOMINAL_F, NOMINAL_COMP)
    kpis = product_kpis(r)
    study_B.append({"P_bar": Plbl, "results": r, "kpis": kpis})
    ch4_f = kpis.get("S-24",      {}).get("F_molh", 0)
    eth_f = kpis.get("Ethane",    {}).get("F_molh", 0)
    btm_f = kpis.get("Deeth_Btm", {}).get("F_molh", 0)
    print(f"    P={Plbl:.2f} bar: S-24={ch4_f:.1f}  Eth={eth_f:.1f}  Btm={btm_f:.1f} mol/h")

# ── Study C: Feed Richness (C2+ content scale) ───────────────────────────────
print("  Study C: Feed composition sweep (T=310 K, P=59.78 bar)")
RICH_SCALES  = [0.5, 0.75, 1.0, 1.5, 2.0]
RICH_LABELS  = ["Lean (0.5x)", "Semi-lean (0.75x)", "Nominal (1.0x)",
                "Semi-rich (1.5x)", "Rich (2.0x)"]
study_C = []
for scale, lbl in zip(RICH_SCALES, RICH_LABELS):
    comp = scaled_comp(scale)
    r = run_case(NOMINAL_T, NOMINAL_P, NOMINAL_F, comp)
    kpis = product_kpis(r)
    study_C.append({"scale": scale, "label": lbl,
                    "CH4_feed_pct": comp["Methane"] * 100,
                    "C2H6_feed_pct": comp.get("Ethane", 0) * 100,
                    "results": r, "kpis": kpis})
    ch4_f = kpis.get("S-24",      {}).get("F_molh", 0)
    eth_f = kpis.get("Ethane",    {}).get("F_molh", 0)
    btm_f = kpis.get("Deeth_Btm", {}).get("F_molh", 0)
    print(f"    {lbl:25s}: S-24={ch4_f:.1f}  Eth={eth_f:.1f}  Btm={btm_f:.1f} mol/h")

# Restore base case before saving
print("  Restoring base-case feed conditions...")
run_case(NOMINAL_T, NOMINAL_P, NOMINAL_F, NOMINAL_COMP)

# ══════════════════════════════════════════════════════════════════════════════
# [13]  GENERATE EXCEL REPORT
# ══════════════════════════════════════════════════════════════════════════════

print(f"\n[13] Generating Excel report -> {EXCEL_FILE} ...")
os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

# ── Style helpers ────────────────────────────────────────────────────────────
_BLUE_FILL  = PatternFill("solid", fgColor="1F497D")
_TEAL_FILL  = PatternFill("solid", fgColor="17375E")
_GREY_FILL  = PatternFill("solid", fgColor="D9E1F2")
_ALT_FILL   = PatternFill("solid", fgColor="EEF2F8")
_GOLD_FILL  = PatternFill("solid", fgColor="C65911")
_GREEN_FILL = PatternFill("solid", fgColor="375623")
_THIN       = Side(style="thin", color="8496B0")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WHITE_BOLD = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FNT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_BODY_FNT   = Font(name="Calibri", size=10)
_BOLD_FNT   = Font(name="Calibri", bold=True, size=10)
_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT       = Alignment(horizontal="left",   vertical="center")
_NUM2       = "0.00"
_NUM3       = "0.000"
_NUM0       = "#,##0"


def _hdr(ws, row: int, col: int, val, fill=None, font=None,
         align=None, num_fmt=None) -> None:
    cell = ws.cell(row=row, column=col, value=val)
    if fill:    cell.fill    = fill
    if font:    cell.font    = font
    if align:   cell.alignment = align
    if num_fmt: cell.number_format = num_fmt
    cell.border = _BORDER


def _val(ws, row: int, col: int, val, fill=None, num_fmt=None,
         bold=False, align=None) -> None:
    cell = ws.cell(row=row, column=col, value=val)
    if fill:    cell.fill    = fill
    if num_fmt: cell.number_format = num_fmt
    cell.font   = _BOLD_FNT if bold else _BODY_FNT
    cell.alignment = align or _LEFT
    cell.border = _BORDER


def _auto_width(ws, min_w: int = 8, max_w: int = 40) -> None:
    for col in ws.columns:
        cw = min_w
        for cell in col:
            if cell.value:
                cw = max(cw, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col[0].column)].width = cw


wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default empty sheet

# ── Sheet 1: Stream Summary ──────────────────────────────────────────────────
ws1 = wb.create_sheet("Stream Summary")
ws1.freeze_panes = "B3"

# Title
ws1.merge_cells("A1:R1")
t = ws1["A1"]
t.value = f"Natural Gas Separation — Stream Summary   |   Generated {datetime.now():%Y-%m-%d %H:%M}"
t.font  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
t.fill  = _BLUE_FILL
t.alignment = _CENTER
ws1.row_dimensions[1].height = 22

STREAM_COLS = [
    "Stream",
    "T (K)", "T (°C)", "P (bar)", "P (kPa)",
    "F (mol/h)", "F (kmol/h)", "Vapor Frac.",
    "CH4 mol%", "C2H6 mol%", "C3H8 mol%",
    "iC4 mol%", "nC4 mol%",
    "iC5 mol%", "nC5 mol%",
    "nC6 mol%", "nC7 mol%",
    "N2 mol%",
]
for c, hdr in enumerate(STREAM_COLS, 1):
    _hdr(ws1, 2, c, hdr, fill=_TEAL_FILL, font=_HEADER_FNT, align=_CENTER)
ws1.row_dimensions[2].height = 30

COMP_ORDER = ["Methane","Ethane","Propane","Isobutane","n-Butane",
              "Isopentane","n-Pentane","n-Hexane","n-Heptane","Nitrogen"]
row = 3
for sname in ALL_STREAMS:
    r = base_case.get(sname)
    fill = _ALT_FILL if row % 2 == 0 else None
    if r is None or r["T_K"] < 50:
        _val(ws1, row, 1, sname, fill=fill, bold=(sname in PRODUCT_STREAMS))
        row += 1
        continue
    bold = sname in PRODUCT_STREAMS
    comp = r["comp"]
    vals = [
        sname,
        r["T_K"], r["T_K"] - 273.15, r["P_bar"], r["P_bar"] * 100,
        r["F_molh"], r["F_molh"] / 1000, r["VF"],
    ] + [comp.get(c, 0) * 100 for c in COMP_ORDER]

    num_fmts = [
        None,
        _NUM2, _NUM2, _NUM3, _NUM2,
        _NUM2, _NUM3, _NUM3,
    ] + [_NUM2] * len(COMP_ORDER)

    for c, (v, nf) in enumerate(zip(vals, num_fmts), 1):
        _val(ws1, row, c, v, fill=fill, num_fmt=nf, bold=bold)

    # Highlight product rows
    if sname in PRODUCT_STREAMS:
        for c in range(1, len(STREAM_COLS) + 1):
            ws1.cell(row=row, column=c).fill = _GREY_FILL

    row += 1

_auto_width(ws1, min_w=9, max_w=14)
ws1.column_dimensions["A"].width = 18

# ── Sheet 2: Equipment & Parameters ─────────────────────────────────────────
ws2 = wb.create_sheet("Equipment")
ws2.merge_cells("A1:G1")
t = ws2["A1"]
t.value = "Unit Operations — Specifications"
t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
t.fill = _BLUE_FILL
t.alignment = _CENTER
ws2.row_dimensions[1].height = 22

EQ_HDRS = ["Tag", "Type", "Parameter 1", "Value 1", "Parameter 2", "Value 2", "Notes"]
for c, h in enumerate(EQ_HDRS, 1):
    _hdr(ws2, 2, c, h, fill=_TEAL_FILL, font=_HEADER_FNT, align=_CENTER)
ws2.row_dimensions[2].height = 28

EQUIP_ROWS = [
    ("COM-01",   "Compressor",        "CalcMode",      "Delta-P",     "DeltaP (Pa)",    202650,  "Eff=0.75"),
    ("SPLIT-01", "Splitter",          "Ratio[0] bypass",f"{15.83/1464:.5f}", "Ratio[1] main", f"{1448/1464:.5f}", "2-outlet"),
    ("COOL-01",  "Cooler",            "CalcMode",      "OutletTemp",  "T_out (K)",      305.3,   "dP=20000 Pa"),
    ("COOL-02",  "Cooler",            "CalcMode",      "OutletTemp",  "T_out (K)",      289.7,   "dP=20000 Pa"),
    ("MIX-01",   "Mixer",             "Inlets",        2,             "-",              "-",     "Bypass + cooled"),
    ("HEX-01",   "HeatExchanger",     "CalcMode",      "SpecifyUA",   "UA (W/K)",       239284,  "U=1000, A=239.28 m2"),
    ("COOL-03",  "Cooler",            "CalcMode",      "OutletTemp",  "T_out (K)",      251.4,   "dP=0 Pa"),
    ("FLASH-01", "Vessel",            "PressureCalc",  "Minimum",     "-",              "-",     ""),
    ("SPLIT-02", "Splitter",          "Ratio[0] HEX-02",f"{483.3/1381:.5f}","Ratio[1] EXP-01",f"{898/1381:.5f}","2-outlet"),
    ("HEX-02",   "HeatExchanger",     "CalcMode",      "SpecifyUA",   "UA (W/K)",       804716,  "U=1000, A=804.72 m2"),
    ("EXP-01",   "Expander",          "CalcMode",      "Delta-P",     "DeltaP (Pa)",    3577572.5,"Eff=0.75"),
    ("FLASH-02", "Vessel",            "PressureCalc",  "Minimum",     "-",              "-",     ""),
    ("VALV-01",  "Valve",             "CalcMode",      "OutletPres",  "P_out (bar)",    26.4133, ""),
    ("VALV-02",  "Valve",             "CalcMode",      "OutletPres",  "P_out (bar)",    27.0083, ""),
    ("MIX-DEMETH","Mixer",            "Inlets",        4,             "-",              "-",     "Combines S-20/21/22/23"),
    ("DEMETH",   "ComponentSeparator","SpecStreamIdx", 0,             "Mode",           "% Inlet MolarFlow","10 comps"),
    ("DEETH",    "ComponentSeparator","SpecStreamIdx", 0,             "Mode",           "% Inlet MolarFlow","10 comps"),
]
for i, row_data in enumerate(EQUIP_ROWS):
    r = i + 3
    fill = _ALT_FILL if i % 2 == 0 else None
    for c, v in enumerate(row_data, 1):
        _val(ws2, r, c, v, fill=fill)

_auto_width(ws2, min_w=10, max_w=28)

# ── Sheet 3: Connections ─────────────────────────────────────────────────────
ws3 = wb.create_sheet("Connections")
ws3.merge_cells("A1:E1")
t = ws3["A1"]
t.value = "Flowsheet Connections"
t.font  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
t.fill  = _BLUE_FILL
t.alignment = _CENTER
ws3.row_dimensions[1].height = 22

CONN_HDRS = ["From Object", "From Port (out)", "To Object", "To Port (in)", "Status"]
for c, h in enumerate(CONN_HDRS, 1):
    _hdr(ws3, 2, c, h, fill=_TEAL_FILL, font=_HEADER_FNT, align=_CENTER)
ws3.row_dimensions[2].height = 28

for i, cdict in enumerate(_CONNECTIONS):
    r = i + 3
    fill = _ALT_FILL if i % 2 == 0 else None
    status = "Connected" if cdict["ok"] else "FAILED"
    sfill  = fill
    sfont  = _BODY_FNT
    if not cdict["ok"]:
        sfill = PatternFill("solid", fgColor="FFD7D7")
        sfont = Font(name="Calibri", size=10, color="9C0006")
    for c, v in enumerate([
        cdict["src"], cdict["src_port"],
        cdict["tgt"], cdict["tgt_port"],
        status,
    ], 1):
        cell = ws3.cell(row=r, column=c, value=v)
        cell.fill      = sfill or PatternFill()
        cell.font      = sfont
        cell.alignment = _CENTER if c in (2, 4, 5) else _LEFT
        cell.border    = _BORDER

_auto_width(ws3, min_w=10, max_w=24)

# ── Sheet 4: Dynamic Study A — Temperature ──────────────────────────────────
def _study_sheet(wb, sheet_name: str, title: str, var_col: str,
                 cases: list, var_key: str) -> None:
    """Generic dynamic-study sheet builder."""
    ws = wb.create_sheet(sheet_name)
    ws.merge_cells(f"A1:N1")
    t = ws["A1"]
    t.value = title
    t.font  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill  = _GOLD_FILL
    t.alignment = _CENTER
    ws.row_dimensions[1].height = 22

    HDRS = [
        var_col,
        "Feed VF", "Feed T (K)", "Feed P (bar)",
        # S-24 = Methane
        "S-24 F (mol/h)", "S-24 CH4%", "S-24 C2H6%",
        # Ethane product
        "Eth F (mol/h)", "Eth CH4%", "Eth C2H6%",
        # Deeth_Btm
        "Btm F (mol/h)", "Btm CH4%", "Btm C3H8%",
        # recovery proxy
        "C2H6 Recovery %",
    ]
    for c, h in enumerate(HDRS, 1):
        _hdr(ws, 2, c, h, fill=PatternFill("solid", fgColor="244062"),
             font=_HEADER_FNT, align=_CENTER)
    ws.row_dimensions[2].height = 36

    # Feed total C2H6 (mol/h) from NOMINAL for recovery calc
    nom_c2h6_feed = NOMINAL_COMP["Ethane"] * NOMINAL_F * 3600.0

    for i, case in enumerate(cases):
        r = i + 3
        fill = _ALT_FILL if i % 2 == 0 else None

        feed_r  = case["results"].get("NG-Feed")
        kpis    = case["kpis"]
        s24     = kpis.get("S-24",      {})
        eth     = kpis.get("Ethane",    {})
        btm     = kpis.get("Deeth_Btm", {})

        feed_vf = feed_r["VF"]   if feed_r else 0
        feed_T  = feed_r["T_K"]  if feed_r else 0
        feed_P  = feed_r["P_bar"]if feed_r else 0

        # C2H6 in Ethane product / C2H6 in feed (approx recovery)
        eth_c2h6_flow = eth.get("F_molh", 0) * eth.get("C2H6_pct", 0) / 100.0
        recovery = eth_c2h6_flow / nom_c2h6_feed * 100.0 if nom_c2h6_feed > 0 else 0

        row_vals = [
            case[var_key],
            feed_vf, feed_T, feed_P,
            s24.get("F_molh", 0), s24.get("CH4_pct", 0), s24.get("C2H6_pct", 0),
            eth.get("F_molh", 0), eth.get("CH4_pct", 0), eth.get("C2H6_pct", 0),
            btm.get("F_molh", 0), btm.get("CH4_pct", 0), btm.get("C3H8_pct", 0),
            recovery,
        ]
        num_fmts = [_NUM2,
                    _NUM3, _NUM2, _NUM3,
                    _NUM2, _NUM2, _NUM2,
                    _NUM2, _NUM2, _NUM2,
                    _NUM2, _NUM2, _NUM2,
                    _NUM2]
        for c, (v, nf) in enumerate(zip(row_vals, num_fmts), 1):
            _val(ws, r, c, v, fill=fill, num_fmt=nf, align=_CENTER)

    _auto_width(ws, min_w=10, max_w=18)


_study_sheet(wb, "Study A - Temperature",
             "Dynamic Study A: Feed Temperature Sweep  (P=59.78 bar, nominal composition)",
             "Feed T (K)",
             [{"T_K": c["T_K"], **c} for c in study_A], "T_K")

_study_sheet(wb, "Study B - Pressure",
             "Dynamic Study B: Feed Pressure Sweep  (T=310 K, nominal composition)",
             "Feed P (bar)",
             [{"P_bar": c["P_bar"], **c} for c in study_B], "P_bar")

# Study C uses a text label → handled slightly differently
ws4 = wb.create_sheet("Study C - Composition")
ws4.merge_cells("A1:O1")
t = ws4["A1"]
t.value = ("Dynamic Study C: Feed Richness (C2+ scaling)  "
           "(T=310 K, P=59.78 bar) — 0.5x=lean, 2.0x=rich")
t.font  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
t.fill  = _GREEN_FILL
t.alignment = _CENTER
ws4.row_dimensions[1].height = 22

HDRS_C = [
    "Case", "C2+ Scale", "Feed CH4%", "Feed C2H6%",
    "S-24 F (mol/h)", "S-24 CH4%",
    "Eth F (mol/h)", "Eth C2H6%",
    "Btm F (mol/h)", "Btm C3H8%",
    "C2H6 Recovery %",
]
for c, h in enumerate(HDRS_C, 1):
    _hdr(ws4, 2, c, h, fill=PatternFill("solid", fgColor="375623"),
         font=_HEADER_FNT, align=_CENTER)
ws4.row_dimensions[2].height = 36

nom_c2h6_feed = NOMINAL_COMP["Ethane"] * NOMINAL_F * 3600.0

for i, case in enumerate(study_C):
    r = i + 3
    fill = _ALT_FILL if i % 2 == 0 else None
    kpis = case["kpis"]
    s24  = kpis.get("S-24",      {})
    eth  = kpis.get("Ethane",    {})
    btm  = kpis.get("Deeth_Btm", {})
    eth_c2h6_flow = eth.get("F_molh", 0) * eth.get("C2H6_pct", 0) / 100.0
    recovery = eth_c2h6_flow / nom_c2h6_feed * 100.0 if nom_c2h6_feed > 0 else 0

    row_vals = [
        case["label"], case["scale"],
        case["CH4_feed_pct"], case["C2H6_feed_pct"],
        s24.get("F_molh", 0), s24.get("CH4_pct", 0),
        eth.get("F_molh", 0), eth.get("C2H6_pct", 0),
        btm.get("F_molh", 0), btm.get("C3H8_pct", 0),
        recovery,
    ]
    num_fmts = [None, _NUM2, _NUM2, _NUM2,
                _NUM2, _NUM2, _NUM2, _NUM2,
                _NUM2, _NUM2, _NUM2]
    for c, (v, nf) in enumerate(zip(row_vals, num_fmts), 1):
        _val(ws4, r, c, v, fill=fill, num_fmt=nf, align=_CENTER)

_auto_width(ws4, min_w=10, max_w=20)

wb.save(EXCEL_FILE)
print(f"  Excel saved -> {EXCEL_FILE}")
print(f"  Sheets: {[s.title for s in wb.worksheets]}")

# ══════════════════════════════════════════════════════════════════════════════
# [14]  CONFIGURE DWSIM DYNAMICS MANAGER
# ══════════════════════════════════════════════════════════════════════════════

print("\n[14] Configuring DWSIM DynamicsManager...")

_DYN_SETUP_ROWS: list[dict] = []   # collected for "Dynamics Setup" Excel sheet

try:
    import pathlib
    import clr
    import System
    from System import TimeSpan, DateTime

    _dm_dll = str(pathlib.Path(dwsim._dwsim_path) / "DWSIM.DynamicsManager.dll")
    clr.AddReference(_dm_dll)
    print(f"  Loaded DynamicsManager DLL: {_dm_dll}")

    from DWSIM.DynamicsManager import (
        Integrator      as DynIntegrator,
        EventSet        as DynEventSet,
        Schedule        as DynSchedule,
        DynamicEvent,
        MonitoredVariable,
    )

    # Enums live in DWSIM.Interfaces as nested types — retrieve via reflection
    # from a prototype event so we don't need a separate import path.
    _ev_proto = DynamicEvent()
    _ev_type   = _ev_proto.GetType()
    from System.Reflection import BindingFlags as _BF
    _PUB = _BF.Public | _BF.Instance
    _ChangeProperty = _ev_type.GetProperty("EventType", _PUB).GetValue(_ev_proto, None)
    _StepChange     = _ev_type.GetProperty("TransitionType", _PUB).GetValue(_ev_proto, None)
    # _ChangeProperty == DynamicsEventType.ChangeProperty (value 0)
    # _StepChange     == DynamicsEventTransitionType.StepChange (value 0)

    # ── feed object UUIDs ─────────────────────────────────────────────────────
    # Use mgr._objects (raw DWSIM objects) to avoid any variable-shadowing from
    # the study loops above (e.g. `s24` gets rebound to a dict in the Excel loop).
    _feed_uuid  = mgr._objects["NG-Feed"].Name
    _s02_uuid   = mgr._objects["S-02"].Name
    _s24_uuid   = mgr._objects["S-24"].Name
    _eth_uuid   = mgr._objects["Ethane"].Name
    _dbtm_uuid  = mgr._objects["Deeth_Btm"].Name

    # ── 1. Integrator ────────────────────────────────────────────────────────
    integrator = DynIntegrator()
    integrator.ID          = "INT-01"
    integrator.Description = "NG Feed Disturbance Study (60s step, 70 min total)"
    integrator.IntegrationStep              = TimeSpan.FromSeconds(60)
    integrator.Duration                     = TimeSpan.FromMinutes(70)
    integrator.ShouldCalculateEquilibrium   = True
    integrator.ShouldCalculatePressureFlow  = False
    integrator.ShouldCalculateControl       = False
    integrator.RealTime                     = False
    integrator.CurrentTime                  = DateTime(2024, 1, 1, 0, 0, 0)

    # ── 1a. Monitored Variables ───────────────────────────────────────────────
    # PROP_MS_0=T(K)  PROP_MS_1=P(Pa)  PROP_MS_2=molar flow (mol/s)
    _mv_defs = [
        (_feed_uuid,  "PROP_MS_0", "NG-Feed Temperature (K)",      250, 380),
        (_feed_uuid,  "PROP_MS_1", "NG-Feed Pressure (Pa)",         40e5, 80e5),
        (_feed_uuid,  "PROP_MS_2", "NG-Feed Molar Flow (mol/s)",    0.2, 0.6),
        (_s02_uuid,   "PROP_MS_0", "S-02 Temperature (K)",          200, 500),
        (_s24_uuid,   "PROP_MS_2", "S-24 Methane Flow (mol/s)",      0.0, 0.5),
        (_s24_uuid,   "PROP_MS_0", "S-24 Temperature (K)",          100, 400),
        (_eth_uuid,   "PROP_MS_2", "Ethane Product Flow (mol/s)",    0.0, 0.1),
        (_dbtm_uuid,  "PROP_MS_2", "Deeth_Btm Flow (mol/s)",         0.0, 0.1),
    ]

    for obj_id, prop_id, desc, lo, hi in _mv_defs:
        mv = MonitoredVariable()
        mv.ObjectID                = obj_id
        mv.PropertyID              = prop_id
        mv.MinimumChartAxisValue   = lo
        mv.MaximumChartAxisValue   = hi
        integrator.MonitoredVariables.Add(mv)
        _DYN_SETUP_ROWS.append({
            "section": "Monitored Variable",
            "name":    desc,
            "detail":  f"ObjectID={obj_id[:8]}…  prop={prop_id}  range=[{lo}, {hi}]",
        })

    _DYN_SETUP_ROWS.insert(0, {
        "section": "Integrator",
        "name":    integrator.ID,
        "detail":  (f"Step=60s  Duration=70min  "
                    f"Equilibrium={integrator.ShouldCalculateEquilibrium}  "
                    f"RealTime={integrator.RealTime}"),
    })
    print(f"  Integrator '{integrator.ID}': step=60s, duration=70min, "
          f"{len(_mv_defs)} monitored vars")

    # ── 2. Event Set ─────────────────────────────────────────────────────────
    event_set = DynEventSet()
    event_set.ID          = "EVT-SET-01"
    event_set.Description = "NG Feed Step Disturbances (T / P / F)"

    # (t_min, prop_id,     value_str,                           description)
    _T_NOM_K = str(NOMINAL_T)           # "310.0"
    _P_NOM_Pa = str(NOMINAL_P)          # "5978180.0"
    _F_NOM    = str(NOMINAL_F)          # mol/s

    _T_UP  = str(NOMINAL_T + 20.0)      # 330 K
    _P_DN  = str(NOMINAL_P - 5e5)       # -5 bar
    _F_UP  = str(NOMINAL_F * 1.10)      # +10 %

    _event_defs = [
        (10, "PROP_MS_0", _T_UP,   "T +20K  (310→330 K)"),
        (20, "PROP_MS_0", _T_NOM_K,"T restore (330→310 K)"),
        (30, "PROP_MS_1", _P_DN,   "P -5 bar (59.78→54.78 bar)"),
        (40, "PROP_MS_1", _P_NOM_Pa,"P restore (54.78→59.78 bar)"),
        (50, "PROP_MS_2", _F_UP,   "F +10%  (1464→1610 mol/h)"),
        (60, "PROP_MS_2", _F_NOM,  "F restore (1610→1464 mol/h)"),
    ]

    for t_min, prop_id, val_str, desc in _event_defs:
        ev = DynamicEvent()
        ev.ID                          = f"EV-{t_min:02d}min"
        ev.Description                 = desc
        ev.TimeStamp                   = DateTime(2024, 1, 1, t_min // 60, t_min % 60, 0)
        ev.EventType                   = _ChangeProperty
        ev.SimulationObjectID          = _feed_uuid
        ev.SimulationObjectProperty    = prop_id
        ev.SimulationObjectPropertyValue = val_str
        ev.SimulationObjectPropertyUnits = ""
        ev.TransitionType              = _StepChange
        ev.Enabled                     = True
        event_set.Events[ev.ID] = ev
        _DYN_SETUP_ROWS.append({
            "section": "Event",
            "name":    ev.ID,
            "detail":  f"t={t_min} min  {prop_id}={val_str}  [{desc}]",
        })

    print(f"  EventSet '{event_set.ID}': {len(_event_defs)} step-change events on NG-Feed")

    # ── 3. Schedule ──────────────────────────────────────────────────────────
    schedule = DynSchedule()
    schedule.ID                    = "SCH-01"
    schedule.Description           = "NG Feed Disturbance Schedule"
    schedule.CurrentIntegrator     = integrator.ID
    schedule.UsesEventList         = True
    schedule.CurrentEventList      = event_set.ID
    schedule.UseCurrentStateAsInitial       = True
    schedule.ResetContentsOfAllObjects      = False
    schedule.UsesCauseAndEffectMatrix       = False

    _DYN_SETUP_ROWS.append({
        "section": "Schedule",
        "name":    schedule.ID,
        "detail":  (f"Integrator={schedule.CurrentIntegrator}  "
                    f"EventList={schedule.CurrentEventList}  "
                    f"UseCurrentState={schedule.UseCurrentStateAsInitial}"),
    })
    print(f"  Schedule '{schedule.ID}': integrator={schedule.CurrentIntegrator}, "
          f"eventList={schedule.CurrentEventList}")

    # ── 4. Register in DynamicsManager ───────────────────────────────────────
    dm = fs.DynamicsManager
    dm.IntegratorList[integrator.ID] = integrator
    dm.EventSetList[event_set.ID]    = event_set
    dm.ScheduleList[schedule.ID]     = schedule
    dm.CurrentSchedule  = schedule.ID
    dm.EnableHistorian  = True
    print(f"  DynamicsManager: historian={dm.EnableHistorian}, "
          f"currentSchedule={dm.CurrentSchedule}")

    # ── 5. Enable Dynamic Mode on the flowsheet ───────────────────────────────
    try:
        fs.DynamicMode = True
        print("  fs.DynamicMode = True")
    except Exception as _e:
        print(f"  WARNING: fs.DynamicMode assignment failed: {_e}")

    try:
        fs.DynamicsManager.ToggleDynamicMode(True)
        print("  ToggleDynamicMode(True) called")
    except Exception as _e:
        print(f"  INFO: ToggleDynamicMode not available (expected): {_e}")

    _dyn_ok = True
    print("  DynamicsManager configuration complete.")

except Exception as _dyn_exc:
    _dyn_ok = False
    print(f"  WARNING: DynamicsManager setup failed: {_dyn_exc}")
    _DYN_SETUP_ROWS.append({
        "section": "ERROR",
        "name":    "Setup failed",
        "detail":  str(_dyn_exc),
    })

# ── Add "Dynamics Setup" Excel sheet ────────────────────────────────────────

print("  Writing 'Dynamics Setup' sheet to Excel...")
ws_dyn = wb.create_sheet("Dynamics Setup")

_TEAL_FILL  = PatternFill("solid", fgColor="1F4E79")
_SECT_FILLS = {
    "Integrator":          PatternFill("solid", fgColor="2E75B6"),
    "Monitored Variable":  PatternFill("solid", fgColor="375623"),
    "Event":               PatternFill("solid", fgColor="7B2C2C"),
    "Schedule":            PatternFill("solid", fgColor="5C3472"),
    "ERROR":               PatternFill("solid", fgColor="C00000"),
}

ws_dyn.merge_cells("A1:C1")
t = ws_dyn["A1"]
t.value     = "DWSIM Dynamics Manager Configuration — NG Feed Disturbance Study"
t.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
t.fill      = _TEAL_FILL
t.alignment = _CENTER
ws_dyn.row_dimensions[1].height = 24

for c, hdr in enumerate(["Section", "ID / Name", "Details"], 1):
    _hdr(ws_dyn, 2, c, hdr, fill=PatternFill("solid", fgColor="2F5496"),
         font=_HEADER_FNT, align=_CENTER)
ws_dyn.row_dimensions[2].height = 30

# Integrator summary block first
_meta_rows = [
    ("Integrator Setting", "IntegrationStep",  "60 seconds"),
    ("Integrator Setting", "Duration",          "70 minutes"),
    ("Integrator Setting", "ShouldCalculateEquilibrium",  "True"),
    ("Integrator Setting", "ShouldCalculatePressureFlow", "False"),
    ("Integrator Setting", "RealTime",           "False"),
    ("Integrator Setting", "CurrentTime",        "2024-01-01 00:00:00"),
    ("Integrator Setting", "EnableHistorian",    "True"),
    ("Schedule Setting",   "UseCurrentStateAsInitial",    "True"),
    ("Schedule Setting",   "ResetContentsOfAllObjects",   "False"),
]
for i, (sec, name, detail) in enumerate(_meta_rows):
    r = i + 3
    fill = _ALT_FILL if i % 2 == 0 else None
    _val(ws_dyn, r, 1, sec,    fill=fill, align=_CENTER)
    _val(ws_dyn, r, 2, name,   fill=fill)
    _val(ws_dyn, r, 3, detail, fill=fill)

# Events + monitored vars
row_offset = 3 + len(_meta_rows)
_val(ws_dyn, row_offset, 1, "", fill=None)  # blank spacer

for i, row in enumerate(_DYN_SETUP_ROWS):
    r = row_offset + 1 + i
    sec    = row["section"]
    sfill  = _SECT_FILLS.get(sec, _ALT_FILL)
    _hdr(ws_dyn, r, 1, sec,           fill=sfill,
         font=Font(name="Calibri", bold=True, size=10, color="FFFFFF"), align=_CENTER)
    _val(ws_dyn, r, 2, row["name"],  fill=_ALT_FILL if i % 2 == 0 else None)
    _val(ws_dyn, r, 3, row["detail"],fill=_ALT_FILL if i % 2 == 0 else None)

ws_dyn.column_dimensions["A"].width = 22
ws_dyn.column_dimensions["B"].width = 36
ws_dyn.column_dimensions["C"].width = 70

wb.save(EXCEL_FILE)
print(f"  Excel updated with 'Dynamics Setup' sheet -> {EXCEL_FILE}")

# ══════════════════════════════════════════════════════════════════════════════
# [15]  SAVE DWSIM FLOWSHEET  (with dynamics embedded)
# ══════════════════════════════════════════════════════════════════════════════

print(f"\n[15] Saving DWSIM flowsheet -> {OUTPUT_FILE} ...")
os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
try:
    dwsim.save_flowsheet(fs, OUTPUT_FILE)
    print(f"  Saved OK -> {OUTPUT_FILE}")
except Exception as exc:
    print(f"  WARNING: DWSIM save failed: {exc}")

print("\n[DONE]")
